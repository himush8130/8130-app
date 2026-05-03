#!/usr/bin/env python3
"""
Generate the worker-facing Excel templates for the 8130 APP initial seed.

Outputs three files into worker_templates/:
  - employees.xlsx
  - employee_availability.xlsx (calendar 01/05/2026 .. 12/07/2026)
  - vehicles.xlsx

Each workbook has:
  * a "data" sheet — what the worker fills
  * a "הוראות" sheet — Hebrew instructions
"""

from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).resolve().parent.parent / "worker_templates"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Common styling
HEADER_FONT = Font(bold=True, size=12)
HEADER_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
SAMPLE_FONT = Font(italic=True, color="888888")
RIGHT_TO_LEFT_VIEW = True

def set_rtl_view(ws):
    if RIGHT_TO_LEFT_VIEW:
        ws.sheet_view.rightToLeft = True

def add_instructions_sheet(wb, lines):
    ws = wb.create_sheet("הוראות")
    set_rtl_view(ws)
    ws.column_dimensions['A'].width = 110
    ws['A1'] = "הוראות מילוי"
    ws['A1'].font = Font(bold=True, size=14)
    for i, line in enumerate(lines, start=3):
        cell = ws.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

def style_header_row(ws, row=1, ncols=None):
    n = ncols or ws.max_column
    for col in range(1, n + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[row].height = 28

def autosize_columns(ws, sample_widths=None):
    """Apply sensible column widths."""
    sample_widths = sample_widths or {}
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if letter in sample_widths:
            ws.column_dimensions[letter].width = sample_widths[letter]

# =====================================================================
# 1. employees.xlsx
# =====================================================================
def build_employees():
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    set_rtl_view(ws)

    headers = ["מספר עובד", "שם", "טלפון", "מקצוע", "הרשאה"]
    ws.append(headers)
    style_header_row(ws)

    sample_rows = [
        [1001, "נועה ברק", "050-1110001", "", "manager"],
        [1002, "אבי כהן", "050-1110002", "", "warehouse"],
        [1003, "יואב לוי", "050-1110003", "רכב", "technician"],
        [1004, "שרון ישראלי", "050-1110004", "חשמל", "technician"],
        [1005, "דנה אברהם", "050-1110005", "אופטיקה", "technician"],
    ]
    for row in sample_rows:
        ws.append(row)
    for r in range(2, 2 + len(sample_rows)):
        for c in range(1, 6):
            ws.cell(row=r, column=c).font = SAMPLE_FONT

    autosize_columns(ws, {"A": 14, "B": 22, "C": 16, "D": 18, "E": 16})

    add_instructions_sheet(wb, [
        "טבלה זו ממלאת את רשימת העובדים במערכת.",
        "כל עובד בשורה נפרדת.",
        "",
        "עמודות:",
        "  • מספר עובד — מספר שלם וייחודי. משמש לכניסה למערכת. דוגמה: 1003.",
        "  • שם — שם מלא בעברית.",
        "  • טלפון — מומלץ בפורמט 050-XXXXXXX. רשות (אבל מומלץ — משמש לכפתורי התקשרות ו-WhatsApp במערכת).",
        "  • מקצוע — שם המקצוע של העובד. רלוונטי בעיקר לטכנאים. אופציונלי. דוגמאות: \"רכב\", \"חשמל\", \"אופטיקה\".",
        "  • הרשאה — אחת משלוש: technician (טכנאי, ברירת מחדל) / warehouse (מחסנאי) / manager (מנהל).",
        "",
        "הערות:",
        "  • השורות שמופיעות באקסל הזה בצבע אפור הן דוגמאות — אפשר למחוק אותן ולהזין נתונים אמיתיים.",
        "  • רשימת המקצועות הסופית תיגזר אוטומטית מהערכים שתכתב בעמודת \"מקצוע\". אין צורך להגדיר מקצועות מראש.",
    ])
    wb.save(OUT_DIR / "employees.xlsx")
    print("✓ employees.xlsx")


# =====================================================================
# 2. employee_availability.xlsx — wide calendar
# =====================================================================
def build_availability():
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    set_rtl_view(ws)

    start = date(2026, 5, 1)
    end   = date(2026, 7, 12)
    days  = []
    cur = start
    while cur <= end:
        days.append(cur)
        cur += timedelta(days=1)

    headers = ["מספר עובד", "שם עובד"] + [d.strftime("%d/%m") for d in days]
    ws.append(headers)
    style_header_row(ws)

    # Sample rows
    sample = [
        [1003, "יואב לוי"],
        [1004, "שרון ישראלי"],
        [1005, "דנה אברהם"],
    ]
    for row in sample:
        # default V for every day
        full_row = row + ["V"] * len(days)
        ws.append(full_row)
        for c in range(1, len(full_row) + 1):
            ws.cell(row=ws.max_row, column=c).font = SAMPLE_FONT

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    for col in range(3, 3 + len(days)):
        ws.column_dimensions[get_column_letter(col)].width = 7

    # Freeze panes after 2 ID columns
    ws.freeze_panes = "C2"

    add_instructions_sheet(wb, [
        "טבלה זו רושמת ימי זמינות / אי-זמינות לכל עובד.",
        "כל עובד בשורה נפרדת. כל יום בטווח 1.5.2026 עד 12.7.2026 הוא עמודה.",
        "",
        "ערך בכל תא:",
        "  • V = העובד זמין באותו יום (ברירת מחדל).",
        "  • X = העובד לא זמין (חופשה / מילואים / מחלה / כל סיבה אחרת).",
        "  • תא ריק יטופל כ-V.",
        "",
        "מילוי מומלץ:",
        "  • לכל העובדים מתוך קובץ employees.xlsx — שורה אחת בקובץ הזה.",
        "  • התחל עם V בכל התאים, וסמן X רק בימים שיודעים מראש שהעובד לא יהיה זמין.",
        "",
        "טווח התאריכים בקובץ זה הוא לטעינה הראשונית בלבד. לאחר מכן, חופשות יתווספו דרך ממשק האפליקציה.",
        "",
        "טיפ: אפשר להקפיא את שתי העמודות הראשונות (View > Freeze Panes) כדי שיישארו גלויות בעת גלילה ימינה.",
    ])
    wb.save(OUT_DIR / "employee_availability.xlsx")
    print("✓ employee_availability.xlsx")


# =====================================================================
# 3. vehicles.xlsx
# =====================================================================
def build_vehicles():
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    set_rtl_view(ws)

    headers = ["מספר רכב", "מקצוע", "מחלקה", "תת מחלקה"]
    ws.append(headers)
    style_header_row(ws)

    # Real data provided by the user (2026-05-03).
    # All entries: profession="רכב", department="רכב".
    real_rows: list[list[str]] = []
    def add_block(sub: str, numbers: list[str]) -> None:
        for n in numbers:
            real_rows.append([n, "רכב", "רכב", sub])

    add_block("סיור",          ["705-164", "705-135", "705-430", "705-476", "705-194", "705-416"])
    add_block("מנהלה",         ["706-677", "706-682", "706-713", "706-714", "706-615", "706-600"])
    add_block("אושקוש סולר",   ["707-156", "707-267"])
    add_block("FMTV",          ["990-084", "990-088", "990-082", "990-405"])
    add_block("אושקוש מרום",   ["684-663"])
    add_block("אמבולנס",       ["561-414"])
    add_block("האמר כתקל",     ["705-285"])
    add_block("האמר בקש",      ["700-402"])
    add_block("האמר רקש",      ["706-549"])
    add_block("FMTV חלפים",    ["990-919"])
    add_block("ריאו מים",      ["676-741"])

    for row in real_rows:
        ws.append(row)

    autosize_columns(ws, {"A": 14, "B": 12, "C": 12, "D": 18})

    add_instructions_sheet(wb, [
        "טבלה זו רושמת את הרכבים והציוד שמטופלים במערכת.",
        "כל רכב בשורה נפרדת.",
        "",
        "עמודות:",
        "  • מספר רכב — מספר/מזהה ייחודי. דוגמה: \"705-164\" או \"G500-001\".",
        "  • מקצוע — שם המקצוע שאחראי על הרכב. חייב להתאים לאחד הערכים שיופיעו בעמודת \"מקצוע\" ב-employees.xlsx (למשל: \"רכב\", \"חשמל\", \"אופטיקה\").",
        "  • מחלקה — מחלקה ארגונית שאליה שייך הרכב. אופציונלי.",
        "  • תת מחלקה — קטגוריה משנית בתוך המחלקה (למשל \"סיור\", \"מנהלה\", \"FMTV\", \"אמבולנס\"). אופציונלי.",
        "",
        "הערה: סוג הרכב הוא הקובע איזה צוות טכנאים רואה את הקריאות לרכב הזה.",
    ])
    wb.save(OUT_DIR / "vehicles.xlsx")
    print(f"✓ vehicles.xlsx ({len(real_rows)} רכבים)")


if __name__ == "__main__":
    build_employees()
    build_availability()
    build_vehicles()
    print(f"\nאקסלים נשמרו ב: {OUT_DIR}")
