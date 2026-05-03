#!/usr/bin/env python3
"""
One-off update for worker_templates/vehicles.xlsx:

  1. Strip non-digit characters from every vehicle_number (e.g. '705-164' -> '705164').
  2. For the 25 vehicles in the LABEL_MAP, replace sub_department with the
     specific label provided by the user (e.g. row '705-430' was 'סיור',
     becomes 'נחשונים').
  3. Leave any other rows the user has added in place — only normalize
     dashes in their vehicle_number column.

Idempotent: running it again on an already-clean file is a no-op.
"""

from pathlib import Path
import openpyxl

PATH = Path(__file__).resolve().parent.parent / "worker_templates" / "vehicles.xlsx"

# Vehicle number (already cleaned) -> the label the user wants in
# "תת מחלקה". Vehicles that didn't have a label keep the category
# header as their sub_department.
LABEL_MAP = {
    # סיור block
    "705164": "סיור 1",
    "705135": "סיור 2",
    "705430": "נחשונים",
    "705476": "חרמש",
    "705194": "מג״ד",
    "705416": "חרמש",
    # מנהלה block
    "706677": "רפואה",
    "706682": "פלוגה מ",
    "706713": "חרמש 1",
    "706714": "נחשונים",
    "706615": "פלוגה ל",
    "706600": "פלוגה כ",
    # Single-row blocks — keep the category as the sub_department.
    "707156": "אושקוש סולר",
    "707267": "אושקוש סולר",
    "990084": "FMTV",
    "990088": "FMTV",
    "990082": "FMTV",
    "990405": "FMTV",
    "684663": "אושקוש מרום",
    "561414": "אמבולנס",
    "705285": "האמר כתקל",
    "700402": "האמר בקש",
    "706549": "האמר רקש",
    "990919": "FMTV חלפים",
    "676741": "ריאו מים",
}


def strip_dashes(value):
    if value is None:
        return None
    s = str(value).strip()
    return "".join(ch for ch in s if ch.isdigit() or (ch.isalpha() and ch.isascii()))


def main():
    wb = openpyxl.load_workbook(PATH)
    ws = wb["data"]

    headers = [cell.value for cell in ws[1]]
    col_num    = headers.index("מספר רכב") + 1
    col_subdep = headers.index("תת מחלקה") + 1

    cleaned_count = 0
    relabeled_count = 0
    untouched_count = 0

    for r in range(2, ws.max_row + 1):
        raw = ws.cell(row=r, column=col_num).value
        if raw is None or str(raw).strip() == "":
            continue

        cleaned = strip_dashes(raw)
        if cleaned != str(raw):
            ws.cell(row=r, column=col_num).value = cleaned
            cleaned_count += 1

        if cleaned in LABEL_MAP:
            old = ws.cell(row=r, column=col_subdep).value
            new = LABEL_MAP[cleaned]
            if old != new:
                ws.cell(row=r, column=col_subdep).value = new
                relabeled_count += 1
        else:
            untouched_count += 1

    wb.save(PATH)
    print(f"✓ vehicles.xlsx updated:")
    print(f"  • {cleaned_count} מספרי רכב נוקו ממקפים")
    print(f"  • {relabeled_count} שורות עודכנו עם תוויות חדשות בתת-מחלקה")
    print(f"  • {untouched_count} שורות לא היו ב-LABEL_MAP — הושארו כפי שהן (פרט להסרת מקפים)")


if __name__ == "__main__":
    main()
