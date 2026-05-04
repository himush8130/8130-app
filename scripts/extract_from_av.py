#!/usr/bin/env python3
"""
Extract employees + their availability from worker_templates/av.xlsx
into the canonical templates.

Inputs:
  worker_templates/av.xlsx — source spreadsheet (Hebrew, RTL).

Outputs:
  worker_templates/employees.xlsx
  worker_templates/employee_availability.xlsx

av.xlsx layout (current):
  row 1: empty
  row 2: dates header in col 4..   (DD/MM)
  row 3: 'מקצוע' | 'שם מלא' | day-of-week labels
  rows 4..29: 26 employees, ordered freely. Profession column may be
              blank — we look up the detailed profession by name from
              the static NAME_PROFESSION map below (originally built
              from av.xlsx when it had grouped headers).

Availability cells use V/X (case-insensitive). Anything else / empty
is treated as available (V).

Profession mapping — the canonical list our system uses is:
  רכב / טנק / מנהל / מחסנאי / נשק
The detailed professions in av.xlsx are folded into these five.
"""

from datetime import date, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES = ROOT / "worker_templates"
SRC = TEMPLATES / "av.xlsx"
EMPLOYEES_DST = TEMPLATES / "employees.xlsx"
AVAILABILITY_DST = TEMPLATES / "employee_availability.xlsx"

# Project's canonical date range for availability
START = date(2026, 5, 1)
END   = date(2026, 7, 12)

# ----- Profession mapping (detailed -> canonical) -----

PROFESSION_MAP = {
    "קצינים":   "מנהל",
    "מכונאות":  "טנק",
    "בק״ש":     "טנק",
    "חשמל":     "טנק",
    "צריח":     "טנק",
    "חילוץ":    "טנק",
    "רכב":      "רכב",
    "נשק":      "נשק",
    "א.נ.ם":    "מחסנאי",
}

# Permissions implied by canonical profession.
PERMISSIONS_BY_PROFESSION = {
    "מנהל":    "manager",
    "מחסנאי":  "warehouse",
    # everything else (רכב, טנק, נשק) defaults to technician.
}

# ----- Name overrides (preferred display spelling) -----
NAME_OVERRIDE = {
    "גלעד גבאי": "גילעד גבאי",
}

# ----- Real employee numbers (provided 2026-05-04) -----
# Keyed by the name as it appears in av.xlsx (BEFORE override).
NAME_NUMBER = {
    "דור קריקב":         8275186,
    "לאון בניאס":         9999999,
    "בן ברקמן":           8048029,
    "דניס בולבינוב":      7667042,
    "משה בנעזרי":         7600591,
    "ראזי אבו עביד":      8592587,
    "גלעד גבאי":          9014610,
    "מיכאל שבצנקו":       8266875,
    "שחף ניסים":          7548287,
    "אלכסנדר לויצקי":     8840972,
    "קיריל צ׳ורקוב":      8831115,
    "דמיטרי איטקין":      8451889,
    "משה שכטר":           9175267,
    "אלמוג כהן":          8183890,
    "רוני ביאנה":         6891337,
    "אלכסיי גרינברג":     5385987,
    "מקסים פלומבויים":    5108573,
    "ירין דוד":           8190085,
    "פואד אבו עאסי":      6134018,
    "דניאל קורולקוב":     8141927,
    "טל אסטרין":          8169903,
    "משה זלקה":           7414543,
    "יוני זנה":           5763399,
    "ניתאי מלכה":         5754689,
    "סרגיי צירלוב":       7471311,
    "תומר סיגלוב":        7777777,
}

# ----- Real phone numbers (provided 2026-05-04) -----
NAME_PHONE = {
    "דור קריקב":         "505580129",
    "לאון בניאס":         "502880886",
    "בן ברקמן":           "526411633",
    "דניס בולבינוב":      "546458538",
    "משה בנעזרי":         "508080543",
    "ראזי אבו עביד":      "523119258",
    "גלעד גבאי":          "542040911",
    "מיכאל שבצנקו":       "507288003",
    "שחף ניסים":          "504380030",
    "אלכסנדר לויצקי":     "529579147",
    "קיריל צ׳ורקוב":      "1234",
    "דמיטרי איטקין":      "526354188",
    "משה שכטר":           "5678",
    "אלמוג כהן":          "542299927",
    "רוני ביאנה":         "528400558",
    "אלכסיי גרינברג":     "532812908",
    "מקסים פלומבויים":    "546505950",
    "ירין דוד":           "504647481",
    "פואד אבו עאסי":      "543932707",
    "דניאל קורולקוב":     "502798896",
    "טל אסטרין":          "508120175",
    "משה זלקה":           "505168600",
    "יוני זנה":           "508123242",
    "ניתאי מלכה":         "546526336",
    "סרגיי צירלוב":       "526230535",
    "תומר סיגלוב":        "536201993",
}

# ----- Name → detailed profession (built when av.xlsx had headers) -----
# Used as fallback when the current av.xlsx has no profession column.
NAME_PROFESSION = {
    "דור קריקב":         "קצינים",
    "לאון בניאס":         "קצינים",
    "בן ברקמן":           "מכונאות",
    "דניס בולבינוב":      "מכונאות",
    "משה בנעזרי":         "מכונאות",
    "ראזי אבו עביד":      "מכונאות",
    "גלעד גבאי":          "מכונאות",
    "מיכאל שבצנקו":       "מכונאות",
    "שחף ניסים":          "בק״ש",
    "אלכסנדר לויצקי":     "בק״ש",
    "קיריל צ׳ורקוב":      "בק״ש",
    "דמיטרי איטקין":      "חשמל",
    "משה שכטר":           "חשמל",
    "אלמוג כהן":          "צריח",
    "רוני ביאנה":         "חילוץ",
    "אלכסיי גרינברג":     "חילוץ",
    "מקסים פלומבויים":    "רכב",
    "ירין דוד":           "רכב",
    "פואד אבו עאסי":      "רכב",
    "דניאל קורולקוב":     "רכב",
    "טל אסטרין":          "רכב",
    "משה זלקה":           "נשק",
    "יוני זנה":           "נשק",
    "ניתאי מלכה":         "נשק",
    "סרגיי צירלוב":       "א.נ.ם",
    "תומר סיגלוב":        "א.נ.ם",
}


# ----- Availability classifier -----

def classify(value) -> str:
    """Return 'V' (available) or 'X' (unavailable). Case-insensitive."""
    if value is None:
        return "V"
    s = str(value).strip().upper()
    if s == "X":
        return "X"
    # Anything else (V, empty, holiday name, substitution note, etc.)
    return "V"


# ----- Read av.xlsx -----

def read_av():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["גיליון1"]

    # Build date column index: column number -> date
    date_cols = {}
    for c in range(4, ws.max_column + 1):
        h = ws.cell(row=2, column=c).value
        if h is None:
            continue
        try:
            dd, mm = str(h).strip().split("/")
            d = date(2026, int(mm), int(dd))
            date_cols[c] = d
        except Exception:
            pass

    employees = []
    current_detailed = None  # falls back to col 2 header if av.xlsx still groups

    for r in range(4, 30):  # employees on rows 4..29
        prof_cell = ws.cell(row=r, column=2).value
        name_cell = ws.cell(row=r, column=3).value

        if prof_cell:
            current_detailed = str(prof_cell).strip()
        if not name_cell or not str(name_cell).strip():
            continue
        raw_name = str(name_cell).strip()
        # Name in av.xlsx is the lookup key. Display name may differ.
        display_name = NAME_OVERRIDE.get(raw_name, raw_name)

        # Prefer the per-name lookup; fall back to the column-grouped header.
        detailed = NAME_PROFESSION.get(raw_name) or current_detailed
        if detailed is None:
            raise SystemExit(f"לא ידוע מקצוע עבור: {raw_name!r}")

        canonical_prof = PROFESSION_MAP.get(detailed)
        if canonical_prof is None:
            raise SystemExit(f"מקצוע לא ממופה: {detailed!r}")

        if raw_name not in NAME_NUMBER:
            raise SystemExit(f"לא ידוע מספר עובד עבור: {raw_name!r}")

        availability = {}
        for c, d in date_cols.items():
            if d < START or d > END:
                continue
            availability[d] = classify(ws.cell(row=r, column=c).value)

        employees.append({
            "number":            NAME_NUMBER[raw_name],
            "name":              display_name,
            "phone":             NAME_PHONE.get(raw_name),
            "detailed":          detailed,
            "profession":        canonical_prof,
            "availability":      availability,
        })

    employees.sort(key=lambda e: e["number"])
    return employees


# ----- Style helpers -----

HEADER_FONT  = Font(bold=True, size=12)
HEADER_FILL  = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[1].height = 28

def add_instructions(wb, lines):
    ws = wb.create_sheet("הוראות")
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 110
    ws["A1"] = "הוראות מילוי"
    ws["A1"].font = Font(bold=True, size=14)
    for i, line in enumerate(lines, start=3):
        cell = ws.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ----- Write employees.xlsx -----

def write_employees(employees):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    ws.sheet_view.rightToLeft = True

    headers = ["מספר עובד", "שם", "טלפון", "מקצוע", "הרשאה"]
    ws.append(headers)
    style_header(ws, len(headers))

    # Permissions derived from canonical profession.
    for emp in employees:
        phone = emp.get("phone") or f"1234{emp['number']}"
        permissions = PERMISSIONS_BY_PROFESSION.get(emp["profession"], "technician")
        ws.append([emp["number"], emp["name"], phone, emp["profession"], permissions])

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    add_instructions(wb, [
        "טבלת עובדים — חולץ מתוך av.xlsx + מיפוי המקצוע המפורט.",
        "",
        "עמודות:",
        "  • מספר עובד — מספר אישי אמיתי (מ-NAME_NUMBER ב-extract_from_av.py).",
        "  • שם — שם מלא, חולץ מ-av.xlsx (עם NAME_OVERRIDE לאיותים מועדפים).",
        "  • טלפון — מספר טלפון אמיתי (מ-NAME_PHONE), עם נפילה ל-1234<מס׳> עבור עובדים שלא ניתן להם מספר.",
        "  • מקצוע — רכב / טנק / מנהל / מחסנאי / נשק.",
        "  • הרשאה — מנהל / מחסנאי / טכנאי. ברירת מחדל: technician.",
        "",
        "מיפוי מקצועות (מקור → מטרה):",
        "  • קצינים → מנהל   (הרשאה: manager)",
        "  • א.נ.ם   → מחסנאי (הרשאה: warehouse)",
        "  • מכונאות / בק״ש / חשמל / צריח / חילוץ → טנק",
        "  • רכב → רכב",
        "  • נשק → נשק",
    ])

    wb.save(EMPLOYEES_DST)
    print(f"✓ employees.xlsx: {len(employees)} עובדים")


# ----- Write employee_availability.xlsx -----

def write_availability(employees):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    ws.sheet_view.rightToLeft = True

    days = []
    cur = START
    while cur <= END:
        days.append(cur)
        cur += timedelta(days=1)

    headers = ["מספר עובד", "שם עובד"] + [d.strftime("%d/%m") for d in days]
    ws.append(headers)
    style_header(ws, len(headers))

    for emp in employees:
        row = [emp["number"], emp["name"]]
        for d in days:
            row.append(emp["availability"].get(d, "V"))
        ws.append(row)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    for col in range(3, 3 + len(days)):
        ws.column_dimensions[get_column_letter(col)].width = 7
    ws.freeze_panes = "C2"

    add_instructions(wb, [
        "טבלת זמינות — חולצה מתוך av.xlsx (ערכי V/X).",
        "",
        "  • V = זמין באותו יום.",
        "  • X = לא זמין.",
        "  • תא ריק במקור טופל כ-V.",
    ])

    wb.save(AVAILABILITY_DST)
    print(f"✓ employee_availability.xlsx: {len(employees)} עובדים × {len(days)} ימים")


# ----- Main -----

def main():
    if not SRC.exists():
        raise SystemExit(f"לא נמצא קובץ מקור: {SRC}")
    employees = read_av()
    if not employees:
        raise SystemExit("לא נמצאו עובדים ב-av.xlsx")
    write_employees(employees)
    write_availability(employees)
    print()
    print("=== סיכום מקצועות ===")
    by_prof = {}
    for e in employees:
        by_prof.setdefault(e["profession"], []).append(e["name"])
    for prof, names in sorted(by_prof.items()):
        print(f"  {prof}: {len(names)} עובדים — {', '.join(names)}")

    print()
    print("=== סיכום זמינות ===")
    total_v = sum(sum(1 for v in e["availability"].values() if v == "V") for e in employees)
    total_x = sum(sum(1 for v in e["availability"].values() if v == "X") for e in employees)
    print(f"  V (זמין):    {total_v}")
    print(f"  X (לא זמין): {total_x}")


if __name__ == "__main__":
    main()
