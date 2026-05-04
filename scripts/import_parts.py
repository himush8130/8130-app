#!/usr/bin/env python3
"""
Load worker_templates/מלאי מכולות.xlsx into the parts table.

Usage:
    python3 scripts/import_parts.py [--dry-run]

Behavior:
    1. Wipes the parts table (truncating dependent rows in
       call_required_parts and part_withdrawals — both empty in dev).
    2. Reads every row from the inventory file.
    3. Generates a unique synthetic SKU when the worker's value
       collides (e.g. '000000000' appearing 12 times → 000000000-001..012).
       The original value is preserved in original_sku for display.
    4. Bulk inserts via the Supabase REST API using the service-role key.

Mapping:
    Excel "שם פריט"             → name
    Excel "מקט"                  → original_sku  (and sku, with -NNN suffix on dups)
    Excel "כמות"                 → quantity
    Excel "מחסן"                 → warehouse
    Excel "ארון"                 → cabinet
    Excel "סוג מאחסן"             → storage_type
    Excel "מספר מאחסן"           → storage_number
    Excel "מספר תא"              → cell_number
    Excel "כמות מינימלית לפריט" → min_threshold
    Excel "פריט בתמורה"           → is_exchange (כן→true, לא→false)
    Excel "ספירת מלאי"           → stock_count
"""

from __future__ import annotations
import argparse
import os
import sys
from collections import Counter
from pathlib import Path
import openpyxl
import urllib.request
import urllib.error
import json

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "worker_templates" / "מלאי מכולות.xlsx"

# ----- env helpers -----

def load_env() -> dict[str, str]:
    env: dict[str, str] = dict(os.environ)
    path = ROOT / ".env.local"
    if path.exists():
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            env.setdefault(k.strip(), v.strip())
    return env


# ----- Excel parsing -----

def to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None

def to_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None

def parse_inventory():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["מלאי"]

    rows = []
    for r in range(2, ws.max_row + 1):
        name = to_text(ws.cell(row=r, column=1).value)
        if not name:
            continue
        original_sku = to_text(ws.cell(row=r, column=2).value)
        quantity   = to_int(ws.cell(row=r, column=3).value) or 0
        warehouse  = to_text(ws.cell(row=r, column=4).value)
        cabinet    = to_int(ws.cell(row=r, column=5).value)
        st_type    = to_text(ws.cell(row=r, column=6).value)
        st_num     = to_int(ws.cell(row=r, column=7).value)
        cell_num   = to_int(ws.cell(row=r, column=8).value)
        min_thresh = to_int(ws.cell(row=r, column=9).value) or 0
        is_exch_v  = to_text(ws.cell(row=r, column=10).value)
        stock_cnt  = to_int(ws.cell(row=r, column=11).value) or 0

        is_exchange = is_exch_v == "כן"

        rows.append({
            "name":            name,
            "original_sku":    original_sku,
            "quantity":        quantity,
            "warehouse":       warehouse,
            "cabinet":         cabinet,
            "storage_type":    st_type,
            "storage_number":  st_num,
            "cell_number":     cell_num,
            "min_threshold":   min_thresh,
            "is_exchange":     is_exchange,
            "stock_count":     stock_cnt,
        })
    return rows


# ----- Synthetic SKU dedup -----

def assign_synthetic_skus(rows):
    """Mutate rows in place adding a unique 'sku' field. Format:
    if a SKU is unique, sku == original_sku.
    if duplicated, suffix '-NNN' to each occurrence in stable order."""
    counts = Counter(r["original_sku"] or "(no-sku)" for r in rows)
    next_idx: Counter[str] = Counter()
    for r in rows:
        key = r["original_sku"] or "(no-sku)"
        if counts[key] == 1 and r["original_sku"] is not None:
            r["sku"] = r["original_sku"]
        else:
            next_idx[key] += 1
            base = r["original_sku"] or "NOSKU"
            r["sku"] = f"{base}-{next_idx[key]:03d}"


# ----- Supabase REST helpers -----

def supabase_request(env, method: str, path: str, body=None):
    url = env["VITE_SUPABASE_URL"].rstrip("/") + "/rest/v1/" + path.lstrip("/")
    key = env["SUPABASE_SERVICE_ROLE_KEY"]
    data_bytes = json.dumps(body).encode("utf-8") if body is not None else None
    req = urllib.request.Request(url, data=data_bytes, method=method)
    req.add_header("apikey", key)
    req.add_header("Authorization", f"Bearer {key}")
    if data_bytes is not None:
        req.add_header("Content-Type", "application/json")
        req.add_header("Prefer", "return=minimal")
    try:
        with urllib.request.urlopen(req) as resp:
            payload = resp.read().decode("utf-8")
            return resp.status, payload
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8")


def truncate_dependents(env):
    # call_required_parts and part_withdrawals reference parts.sku.
    # We MUST clear them before truncating parts, else FK rejects.
    for table in ("call_required_parts", "part_withdrawals"):
        status, body = supabase_request(env, "DELETE", f"{table}?id=neq.00000000-0000-0000-0000-000000000000")
        print(f"  delete from {table}: HTTP {status}")
        if status >= 400:
            raise SystemExit(f"failed to clear {table}: {body}")


def truncate_parts(env):
    status, body = supabase_request(env, "DELETE", "parts?sku=neq.__never__")
    print(f"  delete from parts: HTTP {status}")
    if status >= 400:
        raise SystemExit(f"failed to clear parts: {body}")


def insert_parts(env, rows):
    """POST in batches of 500."""
    BATCH = 500
    inserted = 0
    for i in range(0, len(rows), BATCH):
        chunk = rows[i:i + BATCH]
        status, body = supabase_request(env, "POST", "parts", chunk)
        if status >= 400:
            raise SystemExit(f"insert failed at batch starting {i}: HTTP {status}\n{body}")
        inserted += len(chunk)
        print(f"  inserted {inserted}/{len(rows)}")
    return inserted


# ----- main -----

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="parse and report only")
    args = parser.parse_args()

    if not SRC.exists():
        raise SystemExit(f"missing source: {SRC}")

    rows = parse_inventory()
    print(f"קלט: {len(rows)} שורות מהאקסל")
    counts = Counter(r["original_sku"] or "(no-sku)" for r in rows)
    dups = {k: v for k, v in counts.items() if v > 1}
    print(f"  שכפולי SKU: {len(dups)} מק״טים שונים, סה״כ {sum(dups.values())} שורות.")
    print(f"  דוגמאות: {list(dups.items())[:5]}")

    assign_synthetic_skus(rows)
    print(f"  לאחר ייחודיות: {len(set(r['sku'] for r in rows))} SKU ייחודיים")

    if args.dry_run:
        print("\n(dry-run — לא נשלח כלום ל-DB)")
        return

    env = load_env()
    if not env.get("VITE_SUPABASE_URL") or not env.get("SUPABASE_SERVICE_ROLE_KEY"):
        raise SystemExit("חסרים VITE_SUPABASE_URL או SUPABASE_SERVICE_ROLE_KEY ב-.env.local")

    print("\n--- ניקוי טבלאות תלויות ---")
    truncate_dependents(env)
    print("--- ניקוי טבלת parts ---")
    truncate_parts(env)
    print("--- הזרקת רשומות ל-parts ---")
    inserted = insert_parts(env, rows)
    print(f"\n✓ סיים: {inserted} שורות נוצרו ב-parts.")


if __name__ == "__main__":
    main()
