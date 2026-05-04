#!/usr/bin/env python3
"""
Wipe & reload the lookup tables from worker_templates/.

Loads in order:
    professions (derived from employees+vehicles unique values)
    employees   (from employees.xlsx)
    vehicles    (from vehicles.xlsx)
    employee_availability (from employee_availability.xlsx — only X cells)

Wipes runtime tables that reference employees/vehicles before the reload:
    feedback_notes, call_comments, call_required_parts,
    part_withdrawals, service_calls, employee_availability.

Leaves the parts table alone (already loaded by import_parts.py).

Usage:
    python3 scripts/import_all.py [--dry-run]
"""

from __future__ import annotations
import argparse
import json
import os
import urllib.error
import urllib.request
from datetime import date
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
TEMPLATES = ROOT / "worker_templates"


# ----- env helpers -----

def load_env() -> dict[str, str]:
    env: dict[str, str] = dict(os.environ)
    p = ROOT / ".env.local"
    if p.exists():
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            env.setdefault(k.strip(), v.strip())
    return env


# ----- Supabase REST -----

def sb(env, method: str, path: str, body=None):
    url = env["VITE_SUPABASE_URL"].rstrip("/") + "/rest/v1/" + path.lstrip("/")
    key = env["SUPABASE_SERVICE_ROLE_KEY"]
    data_bytes = json.dumps(body, ensure_ascii=False).encode("utf-8") if body is not None else None
    req = urllib.request.Request(url, data=data_bytes, method=method)
    req.add_header("apikey", key)
    req.add_header("Authorization", f"Bearer {key}")
    if data_bytes is not None:
        req.add_header("Content-Type", "application/json")
        req.add_header("Prefer", "return=minimal")
    try:
        with urllib.request.urlopen(req) as resp:
            return resp.status, resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8")


def must_ok(label: str, result: tuple[int, str]):
    status, body = result
    if status >= 400:
        raise SystemExit(f"{label} failed: HTTP {status}\n{body}")
    return body


# ----- Wipe in FK-safe order -----

def wipe(env):
    print("--- ניקוי טבלאות תלויות ---")
    # IDs are uuid → use a cheap "match anything" filter.
    for table in (
        "feedback_notes",
        "call_comments",
        "call_required_parts",
        "part_withdrawals",
        "service_calls",
        "employee_availability",
    ):
        # employee_availability has composite PK (employee_number, date), no id column
        if table == "employee_availability":
            must_ok(table, sb(env, "DELETE", f"{table}?employee_number=gt.0"))
        else:
            must_ok(table, sb(env, "DELETE", f"{table}?id=neq.00000000-0000-0000-0000-000000000000"))
        print(f"  ✓ {table}")
    # Vehicles & employees use non-uuid PKs.
    must_ok("vehicles",   sb(env, "DELETE", "vehicles?vehicle_number=neq.__never__"))
    print(f"  ✓ vehicles")
    must_ok("employees",  sb(env, "DELETE", "employees?employee_number=gt.0"))
    print(f"  ✓ employees")
    must_ok("professions", sb(env, "DELETE", "professions?id=gt.0"))
    print(f"  ✓ professions")


# ----- Excel loaders -----

def to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None

def to_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def to_phone(v):
    # Excel may store phones as numbers, dropping the leading 0. Restore it.
    s = to_text(v)
    if s is None:
        return None
    digits = ''.join(ch for ch in s if ch.isdigit())
    if len(digits) == 9 and digits.startswith('5'):
        return '0' + digits
    return s


def read_employees():
    wb = openpyxl.load_workbook(TEMPLATES / "employees.xlsx", data_only=True)
    ws = wb["data"]
    rows = []
    for r in range(2, ws.max_row + 1):
        emp_num = to_int(ws.cell(row=r, column=1).value)
        if emp_num is None:
            continue
        rows.append({
            "employee_number":  emp_num,
            "name":             to_text(ws.cell(row=r, column=2).value) or "",
            "phone":            to_phone(ws.cell(row=r, column=3).value),
            "profession_name":  to_text(ws.cell(row=r, column=4).value),
            "permissions":      to_text(ws.cell(row=r, column=5).value) or "technician",
        })
    return rows


def read_vehicles():
    wb = openpyxl.load_workbook(TEMPLATES / "vehicles.xlsx", data_only=True)
    ws = wb["data"]
    rows = []
    for r in range(2, ws.max_row + 1):
        veh = to_text(ws.cell(row=r, column=1).value)
        if not veh:
            continue
        rows.append({
            "vehicle_number":  veh,
            "type_name":       to_text(ws.cell(row=r, column=2).value) or "",
            "department":      to_text(ws.cell(row=r, column=3).value),
            "sub_department":  to_text(ws.cell(row=r, column=4).value),
        })
    return rows


def read_availability():
    wb = openpyxl.load_workbook(TEMPLATES / "employee_availability.xlsx", data_only=True)
    ws = wb["data"]

    # Header: col 1=מספר עובד, col 2=שם עובד, cols 3+ = DD/MM
    date_cols: list[tuple[int, date]] = []
    for c in range(3, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is None:
            continue
        try:
            dd, mm = str(h).strip().split("/")
            date_cols.append((c, date(2026, int(mm), int(dd))))
        except Exception:
            pass

    rows = []
    for r in range(2, ws.max_row + 1):
        emp_num = to_int(ws.cell(row=r, column=1).value)
        if emp_num is None:
            continue
        for c, d in date_cols:
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if str(v).strip().upper() == "X":
                rows.append({
                    "employee_number": emp_num,
                    "date":            d.isoformat(),
                })
    return rows


# ----- Insert helpers -----

def insert_batches(env, table: str, rows: list[dict], batch_size: int = 500) -> int:
    inserted = 0
    for i in range(0, len(rows), batch_size):
        chunk = rows[i:i + batch_size]
        must_ok(f"{table} insert @ {i}", sb(env, "POST", table, chunk))
        inserted += len(chunk)
        print(f"  {table}: {inserted}/{len(rows)}")
    return inserted


# ----- main -----

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    employees = read_employees()
    vehicles  = read_vehicles()
    avail     = read_availability()

    profession_names = sorted({
        e["profession_name"] for e in employees if e["profession_name"]
    } | {
        v["type_name"] for v in vehicles if v["type_name"]
    })
    professions = [{"name": n} for n in profession_names]

    print(f"קלט מהאקסלים:")
    print(f"  professions:           {len(professions)} ({', '.join(profession_names)})")
    print(f"  employees:             {len(employees)}")
    print(f"  vehicles:              {len(vehicles)}")
    print(f"  employee_availability: {len(avail)} (תאי X)")

    if args.dry_run:
        print("\n(dry-run — לא נוגעים ב-DB)")
        return

    env = load_env()
    if not env.get("VITE_SUPABASE_URL") or not env.get("SUPABASE_SERVICE_ROLE_KEY"):
        raise SystemExit("חסרים VITE_SUPABASE_URL או SUPABASE_SERVICE_ROLE_KEY ב-.env.local")

    wipe(env)

    print("\n--- טעינה מחדש ---")
    insert_batches(env, "professions",            professions)
    insert_batches(env, "employees",              employees)
    insert_batches(env, "vehicles",               vehicles)
    insert_batches(env, "employee_availability",  avail)

    print("\n✓ הסתיים בהצלחה")


if __name__ == "__main__":
    main()
